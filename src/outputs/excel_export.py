"""Excel Export Module - Professional ROI Report Generation"""

from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """
    Generate professional Excel reports for ROI presentations.

    Creates multi-sheet workbooks with:
    - Executive Summary
    - ROI Calculation Detail
    - TCO Comparison
    - Sensitivity Analysis
    - Assumptions
    """

    # Color scheme
    COLORS = {
        "header_bg": "366092",
        "header_text": "FFFFFF",
        "accent": "4472C4",
        "positive": "70AD47",
        "negative": "C00000",
        "neutral": "FFC000",
        "light_gray": "F2F2F2",
        "border": "D9D9D9",
    }

    def __init__(self, config_path: Optional[str] = None):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install: pip install openpyxl")

        self.config = self._load_config(config_path)
        self._setup_styles()

    def _load_config(self, path: Optional[str]) -> dict:
        if path and Path(path).exists():
            with open(path, "r") as f:
                return json.load(f).get("excel_settings", {})
        return {
            "company_name": "Enterprise Solutions",
            "header_color": self.COLORS["header_bg"],
            "accent_color": self.COLORS["accent"],
        }

    def _setup_styles(self):
        """Setup reusable styles."""
        self.header_font = Font(bold=True, size=11, color=self.COLORS["header_text"])
        self.header_fill = PatternFill(
            start_color=self.config.get("header_color", self.COLORS["header_bg"]),
            end_color=self.config.get("header_color", self.COLORS["header_bg"]),
            fill_type="solid"
        )
        self.title_font = Font(bold=True, size=14, color=self.COLORS["header_bg"])
        self.subtitle_font = Font(bold=True, size=12, color=self.COLORS["accent"])
        self.money_format = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'
        self.percent_format = '0.0%'
        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS["border"]),
            right=Side(style='thin', color=self.COLORS["border"]),
            top=Side(style='thin', color=self.COLORS["border"]),
            bottom=Side(style='thin', color=self.COLORS["border"])
        )

    def _apply_header_row(self, ws, row: int, headers: List[str], start_col: int = 1):
        """Apply header formatting to a row."""
        for i, header in enumerate(headers, start_col):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border

    def _format_currency(self, ws, cell_ref: str, value: float):
        """Format a cell as currency."""
        ws[cell_ref] = value
        ws[cell_ref].number_format = self.money_format
        ws[cell_ref].alignment = Alignment(horizontal='right')

    def _format_percent(self, ws, cell_ref: str, value: float):
        """Format a cell as percentage."""
        ws[cell_ref] = value / 100 if value > 1 else value
        ws[cell_ref].number_format = self.percent_format
        ws[cell_ref].alignment = Alignment(horizontal='right')

    def generate_report(
        self,
        prospect_name: str = "Prospect",
        inputs: Optional[Dict] = None,
        roi_result: Optional[Any] = None,
        tco_result: Optional[Any] = None,
        sensitivity_results: Optional[List] = None,
        output_path: str = "output/roi_report.xlsx"
    ) -> str:
        """
        Generate complete ROI report workbook.

        Args:
            prospect_name: Client/prospect name
            inputs: Input assumptions dictionary
            roi_result: ROIResult object or dict
            tco_result: TCOResult object or dict
            sensitivity_results: List of sensitivity analysis results
            output_path: Output file path

        Returns:
            Path to generated file
        """
        wb = Workbook()
        inputs = inputs or {}

        # Convert result objects to dicts if needed
        roi_data = roi_result.to_dict() if hasattr(roi_result, 'to_dict') else (roi_result or {})
        tco_data = tco_result.to_dict() if hasattr(tco_result, 'to_dict') else (tco_result or {})

        # Create sheets
        self._create_executive_summary(wb, prospect_name, inputs, roi_data)
        self._create_roi_detail(wb, inputs, roi_data)
        self._create_tco_comparison(wb, tco_data)
        self._create_sensitivity_sheet(wb, sensitivity_results)
        self._create_assumptions_sheet(wb, inputs)

        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _create_executive_summary(self, wb, prospect_name: str, inputs: dict, roi_data: dict):
        """Create Executive Summary sheet."""
        ws = wb.active
        ws.title = "Executive Summary"

        # Title
        ws['A1'] = f"ROI Analysis: {prospect_name}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Prepared by: {self.config.get('company_name', 'Enterprise Solutions')}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'].font = Font(italic=True, size=10)

        # Key Metrics Section
        ws['A5'] = "KEY METRICS"
        ws['A5'].font = self.subtitle_font

        self._apply_header_row(ws, 6, ["Metric", "Value", "Assessment"], 1)

        metrics = [
            ("3-Year ROI", f"{roi_data.get('roi_percent', 0):.1f}%",
             "Strong" if roi_data.get('roi_percent', 0) > 100 else "Moderate"),
            ("Payback Period", f"{roi_data.get('payback_months', 0):.0f} months",
             "Quick" if roi_data.get('payback_months', 0) < 18 else "Standard"),
            ("Annual Savings", f"${roi_data.get('annual_savings', 0):,.0f}", ""),
            ("Net Present Value", f"${roi_data.get('npv', 0):,.0f}",
             "Positive" if roi_data.get('npv', 0) > 0 else "Review"),
            ("Total Investment", f"${roi_data.get('total_investment', 0):,.0f}", ""),
        ]

        for i, (metric, value, assessment) in enumerate(metrics, 7):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'C{i}'] = assessment
            for col in ['A', 'B', 'C']:
                ws[f'{col}{i}'].border = self.thin_border

            # Color code assessments
            if assessment in ["Strong", "Quick", "Positive"]:
                ws[f'C{i}'].font = Font(color=self.COLORS["positive"], bold=True)
            elif assessment == "Review":
                ws[f'C{i}'].font = Font(color=self.COLORS["negative"], bold=True)

        # Investment Summary
        ws['A14'] = "INVESTMENT SUMMARY"
        ws['A14'].font = self.subtitle_font

        summary_items = [
            ("Implementation Cost", roi_data.get('implementation_cost', 0)),
            ("Annual License (x3)", roi_data.get('annual_license', 0) * 3),
            ("Total 3-Year Investment", roi_data.get('total_investment', 0)),
            ("Total 3-Year Benefits", sum(roi_data.get('yearly_benefits', [0]))),
            ("Net 3-Year Benefit", roi_data.get('three_year_net_benefit', 0)),
        ]

        for i, (label, value) in enumerate(summary_items, 15):
            ws[f'A{i}'] = label
            self._format_currency(ws, f'B{i}', value)
            ws[f'A{i}'].border = self.thin_border
            ws[f'B{i}'].border = self.thin_border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15

    def _create_roi_detail(self, wb, inputs: dict, roi_data: dict):
        """Create ROI Calculation Detail sheet."""
        ws = wb.create_sheet("ROI Detail")

        ws['A1'] = "ROI CALCULATION DETAIL"
        ws['A1'].font = self.title_font

        # Yearly breakdown
        ws['A3'] = "Year-by-Year Analysis"
        ws['A3'].font = self.subtitle_font

        self._apply_header_row(ws, 4, ["Year", "Benefits", "Costs", "Net Cash Flow", "Cumulative"], 1)

        yearly_benefits = roi_data.get('yearly_benefits', [0, 0, 0, 0])
        yearly_costs = roi_data.get('yearly_costs', [0, 0, 0, 0])
        yearly_cf = roi_data.get('yearly_cash_flows', [0, 0, 0, 0])

        cumulative = 0
        for i, (year_idx, benefit, cost, cf) in enumerate(
            zip(range(len(yearly_benefits)), yearly_benefits, yearly_costs, yearly_cf), 5
        ):
            cumulative += cf
            ws[f'A{i}'] = f"Year {year_idx}" if year_idx > 0 else "Year 0 (Implementation)"
            self._format_currency(ws, f'B{i}', benefit)
            self._format_currency(ws, f'C{i}', cost)
            self._format_currency(ws, f'D{i}', cf)
            self._format_currency(ws, f'E{i}', cumulative)

            # Highlight negative cash flows
            if cf < 0:
                ws[f'D{i}'].font = Font(color=self.COLORS["negative"])

            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{i}'].border = self.thin_border

        # Add bar chart for yearly benefits vs costs
        if len(yearly_benefits) > 1:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Benefits vs Costs by Year"
            chart.y_axis.title = "Amount ($)"

            data_end_row = 4 + len(yearly_benefits)
            benefits_ref = Reference(ws, min_col=2, min_row=4, max_row=data_end_row)
            costs_ref = Reference(ws, min_col=3, min_row=4, max_row=data_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=5, max_row=data_end_row)

            chart.add_data(benefits_ref, titles_from_data=True)
            chart.add_data(costs_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            ws.add_chart(chart, "G3")

    def _create_tco_comparison(self, wb, tco_data: dict):
        """Create TCO Comparison sheet."""
        ws = wb.create_sheet("TCO Comparison")

        ws['A1'] = "TOTAL COST OF OWNERSHIP COMPARISON"
        ws['A1'].font = self.title_font

        if not tco_data:
            ws['A3'] = "No TCO data available"
            return

        # Summary metrics
        ws['A3'] = "Summary"
        ws['A3'].font = self.subtitle_font

        self._apply_header_row(ws, 4, ["Metric", "Current State", "Future State", "Savings"], 1)

        current_tco = tco_data.get('current_state_tco', 0)
        future_tco = tco_data.get('future_state_tco', 0)
        savings = tco_data.get('tco_savings', 0)
        savings_pct = tco_data.get('savings_percent', 0)

        ws['A5'] = f"{tco_data.get('years', 5)}-Year TCO"
        self._format_currency(ws, 'B5', current_tco)
        self._format_currency(ws, 'C5', future_tco)
        self._format_currency(ws, 'D5', savings)

        ws['A6'] = "Savings %"
        ws['B6'] = "-"
        ws['C6'] = "-"
        self._format_percent(ws, 'D6', savings_pct)

        for row in [5, 6]:
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.thin_border

        # Yearly comparison
        yearly_comp = tco_data.get('yearly_comparison', [])
        if yearly_comp:
            ws['A9'] = "Year-by-Year Comparison"
            ws['A9'].font = self.subtitle_font

            self._apply_header_row(ws, 10, ["Year", "Current (Cumulative)", "Future (Cumulative)", "Savings"], 1)

            for i, year_data in enumerate(yearly_comp, 11):
                ws[f'A{i}'] = f"Year {year_data.get('year', i-11)}"
                self._format_currency(ws, f'B{i}', year_data.get('current_cumulative', 0))
                self._format_currency(ws, f'C{i}', year_data.get('future_cumulative', 0))
                self._format_currency(ws, f'D{i}', year_data.get('cumulative_savings', 0))

                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{i}'].border = self.thin_border

    def _create_sensitivity_sheet(self, wb, sensitivity_results: Optional[List]):
        """Create Sensitivity Analysis sheet."""
        ws = wb.create_sheet("Sensitivity Analysis")

        ws['A1'] = "SENSITIVITY ANALYSIS"
        ws['A1'].font = self.title_font

        if not sensitivity_results:
            ws['A3'] = "No sensitivity analysis data available"
            return

        ws['A3'] = "Variable Impact on ROI"
        ws['A3'].font = self.subtitle_font

        self._apply_header_row(ws, 4, ["Variable", "Low Value", "Base Value", "High Value",
                                        "Low ROI", "Base ROI", "High ROI", "Impact Range"], 1)

        for i, result in enumerate(sensitivity_results, 5):
            if hasattr(result, 'to_dict'):
                data = result.to_dict()
            else:
                data = result

            ws[f'A{i}'] = data.get('variable', '')
            self._format_currency(ws, f'B{i}', data.get('low_value', 0))
            self._format_currency(ws, f'C{i}', data.get('base_value', 0))
            self._format_currency(ws, f'D{i}', data.get('high_value', 0))
            self._format_percent(ws, f'E{i}', data.get('low_roi', 0))
            self._format_percent(ws, f'F{i}', data.get('base_roi', 0))
            self._format_percent(ws, f'G{i}', data.get('high_roi', 0))
            self._format_percent(ws, f'H{i}', data.get('impact_range', 0))

            for col in 'ABCDEFGH':
                ws[f'{col}{i}'].border = self.thin_border

    def _create_assumptions_sheet(self, wb, inputs: dict):
        """Create Assumptions sheet."""
        ws = wb.create_sheet("Assumptions")

        ws['A1'] = "INPUT ASSUMPTIONS"
        ws['A1'].font = self.title_font

        ws['A3'] = "Financial Inputs"
        ws['A3'].font = self.subtitle_font

        self._apply_header_row(ws, 4, ["Parameter", "Value", "Notes"], 1)

        assumptions = [
            ("Current Annual Cost", inputs.get('current_annual_cost', 0), "Baseline operational costs"),
            ("Efficiency Gain", inputs.get('efficiency_gain', 0), "Expected improvement (0-100%)"),
            ("Implementation Cost", inputs.get('implementation_cost', 0), "One-time setup costs"),
            ("Annual License Cost", inputs.get('annual_license', 0), "Recurring subscription"),
            ("Analysis Period", inputs.get('years', 3), "Years for ROI calculation"),
            ("Discount Rate", inputs.get('discount_rate', 0.10), "For NPV calculation"),
        ]

        for i, (param, value, notes) in enumerate(assumptions, 5):
            ws[f'A{i}'] = param
            if isinstance(value, float) and value < 1:
                self._format_percent(ws, f'B{i}', value * 100)
            elif isinstance(value, (int, float)) and value > 100:
                self._format_currency(ws, f'B{i}', value)
            else:
                ws[f'B{i}'] = value
            ws[f'C{i}'] = notes

            for col in ['A', 'B', 'C']:
                ws[f'{col}{i}'].border = self.thin_border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
